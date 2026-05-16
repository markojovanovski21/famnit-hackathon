"""
Car Suspension Tracker
======================
Tracks a point on the car roof and a point on the wheel hub in a side-profile
video, then plots the vertical displacement of the wheel relative to the roof
(i.e. suspension travel).

Requirements:
    pip install opencv-contrib-python numpy matplotlib openpyxl

Usage:
    python track_suspension.py --video your_video.mp4

Controls (selection window):
    - A window will open showing the first frame.
    - You will be prompted TWICE via the terminal:
        1. Draw a box around the ROOF tracking target (press ENTER/SPACE to confirm, C to cancel)
        2. Draw a box around the WHEEL HUB tracking target (same controls)
    - The tracker then runs automatically.
    - Press Q at any time to stop early.

Output:
    - An annotated video  : track_suspension_out.mp4
    - A displacement plot : suspension_displacement.png
    - A CSV of raw data   : suspension_data.csv
"""

import cv2
import numpy as np
import matplotlib.pyplot as plt
import argparse
import os
import sys


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
TRACKER_TYPE   = "CSRT"          # Best accuracy; alternatives: KCF, MIL, MOSSE
PIXEL_PER_MM   = None            # Set to a float (px/mm) to get real-world units.
                                 # None = results reported in pixels.
OUTPUT_VIDEO   = "track_suspension_out.mp4"
OUTPUT_PLOT    = "suspension_displacement.png"
OUTPUT_CSV     = "suspension_data.csv"
MAX_DISPLAY_W  = 1280            # Preview window max width  (pixels)
MAX_DISPLAY_H  = 720             # Preview window max height (pixels)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_tracker():
    """Return a fresh OpenCV tracker, compatible with both old and new OpenCV layouts.

    OpenCV 4.5.1+ moved most trackers into cv2.legacy.*_create().
    We try the modern location first, then fall back to the top-level API.
    """
    t = TRACKER_TYPE.upper()

    # Map tracker name → (modern constructor, legacy constructor)
    TRACKER_MAP = {
        "CSRT":  ("cv2.TrackerCSRT_create",          "cv2.legacy.TrackerCSRT_create"),
        "KCF":   ("cv2.TrackerKCF_create",           "cv2.legacy.TrackerKCF_create"),
        "MIL":   ("cv2.TrackerMIL_create",           "cv2.legacy.TrackerMIL_create"),
        "MOSSE": ("cv2.legacy.TrackerMOSSE_create",  "cv2.legacy.TrackerMOSSE_create"),
    }

    if t not in TRACKER_MAP:
        raise ValueError(f"Unknown tracker type: {TRACKER_TYPE}. "
                         f"Choose from: {list(TRACKER_MAP)}")

    modern_path, legacy_path = TRACKER_MAP[t]

    for attr_path in (modern_path, legacy_path):
        try:
            parts = attr_path.split(".")          # e.g. ["cv2", "legacy", "TrackerCSRT_create"]
            obj = sys.modules["cv2"]
            for part in parts[1:]:                # skip "cv2"
                obj = getattr(obj, part)
            return obj()
        except AttributeError:
            continue

    raise RuntimeError(
        f"Could not create tracker '{t}'. "
        "Make sure opencv-contrib-python is installed:\n"
        "  pip install opencv-contrib-python"
    )


def bbox_center(bbox):
    """Return (cx, cy) float centre of an OpenCV bounding box (x,y,w,h)."""
    x, y, w, h = bbox
    return x + w / 2.0, y + h / 2.0


def select_roi_with_label(frame, label):
    """Show frame and ask user to draw an ROI. Returns (x, y, w, h)."""
    print(f"\n>>> Draw a bounding box around the {label}.")
    print("    Drag with mouse, then press ENTER or SPACE to confirm (C to redo).")
    roi = cv2.selectROI(f"Select: {label}", frame, fromCenter=False, showCrosshair=True)
    cv2.destroyWindow(f"Select: {label}")
    if roi == (0, 0, 0, 0):
        print(f"No ROI selected for {label}. Exiting.")
        sys.exit(1)
    return roi


def preview_scale(frame):
    """Return a display-safe copy scaled to fit MAX_DISPLAY_W x MAX_DISPLAY_H.
    The original frame is never modified; the output video uses full resolution."""
    h, w = frame.shape[:2]
    scale = min(MAX_DISPLAY_W / w, MAX_DISPLAY_H / h, 1.0)  # never upscale
    if scale == 1.0:
        return frame, 1.0
    new_w, new_h = int(w * scale), int(h * scale)
    return cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_AREA), scale


def select_roi_scaled(frame, label):
    """Show a screen-safe scaled version for ROI selection; return coords in
    original (full-resolution) pixel space."""
    small, scale = preview_scale(frame)
    print(f"\n>>> Draw a bounding box around the {label}.")
    print("    Drag with mouse, then press ENTER or SPACE to confirm (C to redo).")
    roi_s = cv2.selectROI(f"Select: {label}", small, fromCenter=False, showCrosshair=True)
    cv2.destroyWindow(f"Select: {label}")
    if roi_s == (0, 0, 0, 0):
        print(f"No ROI selected for {label}. Exiting.")
        sys.exit(1)
    # Map back to full-resolution coordinates
    x, y, w, h = roi_s
    return (int(x / scale), int(y / scale), int(w / scale), int(h / scale))


def draw_crosshair(img, cx, cy, color, size=10, thickness=2):
    cx, cy = int(cx), int(cy)
    cv2.line(img, (cx - size, cy), (cx + size, cy), color, thickness)
    cv2.line(img, (cx, cy - size), (cx, cy + size), color, thickness)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(video_path: str):
    if not os.path.isfile(video_path):
        print(f"Error: file not found: {video_path}")
        sys.exit(1)

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print("Error: cannot open video.")
        sys.exit(1)

    fps    = cap.get(cv2.CAP_PROP_FPS) or 30.0
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total  = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    print(f"Video: {width}x{height}  {fps:.1f} fps  {total} frames")

    # --- Read first frame ---
    ret, first_frame = cap.read()
    if not ret:
        print("Error: cannot read first frame.")
        sys.exit(1)

    # --- Select ROIs (shown in a screen-safe scaled window) ---
    roof_bbox  = select_roi_scaled(first_frame, "CAR ROOF (body reference)")
    wheel_bbox = select_roi_scaled(first_frame, "WHEEL HUB (suspension point)")

    # --- Initialise trackers ---
    tracker_roof  = make_tracker()
    tracker_wheel = make_tracker()
    tracker_roof.init(first_frame,  roof_bbox)
    tracker_wheel.init(first_frame, wheel_bbox)

    # --- Video writer ---
    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
    writer = cv2.VideoWriter(OUTPUT_VIDEO, fourcc, fps, (width, height))

    # --- Data storage ---
    times          = []   # seconds
    roof_ys        = []   # vertical pixel position of roof centre
    wheel_ys       = []   # vertical pixel position of wheel centre
    rel_displacements = []  # wheel_y - roof_y  (positive = wheel moved down)

    roof_cx0, roof_cy0   = bbox_center(roof_bbox)
    wheel_cx0, wheel_cy0 = bbox_center(wheel_bbox)
    baseline_rel_y = wheel_cy0 - roof_cy0   # initial offset (subtract to get Δ)

    frame_idx = 0
    lost_roof  = False
    lost_wheel = False

    cap.set(cv2.CAP_PROP_POS_FRAMES, 0)   # rewind

    print("\nTracking … press Q in the preview window to stop early.\n")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        t = frame_idx / fps

        # --- Update trackers ---
        ok_roof,  bbox_roof  = tracker_roof.update(frame)
        ok_wheel, bbox_wheel = tracker_wheel.update(frame)

        vis = frame.copy()

        if ok_roof:
            rcx, rcy = bbox_center(bbox_roof)
            x, y, w, h = [int(v) for v in bbox_roof]
            cv2.rectangle(vis, (x, y), (x+w, y+h), (0, 255, 0), 2)
            draw_crosshair(vis, rcx, rcy, (0, 255, 0))
            roof_ys.append(rcy)
        else:
            if not lost_roof:
                print(f"  WARNING: roof tracker lost at frame {frame_idx} ({t:.2f}s)")
                lost_roof = True
            roof_ys.append(np.nan)
            cv2.putText(vis, "ROOF LOST", (10, 60),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255), 2)

        if ok_wheel:
            wcx, wcy = bbox_center(bbox_wheel)
            x, y, w, h = [int(v) for v in bbox_wheel]
            cv2.rectangle(vis, (x, y), (x+w, y+h), (0, 100, 255), 2)
            draw_crosshair(vis, wcx, wcy, (0, 100, 255))
            wheel_ys.append(wcy)
        else:
            if not lost_wheel:
                print(f"  WARNING: wheel tracker lost at frame {frame_idx} ({t:.2f}s)")
                lost_wheel = True
            wheel_ys.append(np.nan)
            cv2.putText(vis, "WHEEL LOST", (10, 90),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255), 2)

        # Relative displacement: positive = wheel dropped relative to body
        if ok_roof and ok_wheel:
            rel = (wcy - rcy) - baseline_rel_y
            if PIXEL_PER_MM:
                rel_mm = rel / PIXEL_PER_MM
                rel_displacements.append(rel_mm)
                label = f"Susp: {rel_mm:+.1f} mm"
            else:
                rel_displacements.append(rel)
                label = f"Susp: {rel:+.1f} px"
        else:
            rel_displacements.append(np.nan)
            label = "Susp: ---"

        times.append(t)

        # HUD
        cv2.putText(vis, label, (10, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2,
                    lineType=cv2.LINE_AA)
        unit = "mm" if PIXEL_PER_MM else "px"
        cv2.putText(vis, f"Frame {frame_idx}/{total}  t={t:.2f}s",
                    (width - 280, 30), cv2.FONT_HERSHEY_SIMPLEX,
                    0.6, (200, 200, 200), 1)

        # Draw vertical line between roof and wheel centres (if both valid)
        if ok_roof and ok_wheel:
            rcx2, rcy2  = bbox_center(bbox_roof)
            wcx2, wcy2  = bbox_center(bbox_wheel)
            cv2.line(vis, (int(rcx2), int(rcy2)), (int(wcx2), int(wcy2)),
                     (255, 255, 0), 1, cv2.LINE_AA)

        writer.write(vis)

        preview, _ = preview_scale(vis)
        cv2.imshow("Suspension Tracker  [Q = quit]", preview)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            print("Stopped early by user.")
            break

        frame_idx += 1

    cap.release()
    writer.release()
    cv2.destroyAllWindows()

    print(f"\nAnnotated video saved → {OUTPUT_VIDEO}")

    # -------------------------------------------------------------------
    # Save Excel (.xlsx)  — data sheet + embedded chart sheet
    # -------------------------------------------------------------------
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    unit = "mm" if PIXEL_PER_MM else "px"
    OUTPUT_XLSX = OUTPUT_CSV.replace(".csv", ".xlsx")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Raw Data ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tracking Data"

    # Header style
    hdr_fill   = PatternFill("solid", fgColor="2C3E50")
    hdr_font   = Font(color="FFFFFF", bold=True)
    hdr_align  = Alignment(horizontal="center")
    thin_side  = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side,
                         bottom=thin_side, top=thin_side)

    headers = ["Frame", "Time (s)",
               f"Roof Y ({unit})", f"Wheel Y ({unit})",
               f"Roof X ({unit})", f"Wheel X ({unit})",
               f"Suspension Travel ({unit})"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 24

    # Alternate row shading
    fill_even = PatternFill("solid", fgColor="EBF5FB")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    for i, (t, ry, wy, rd) in enumerate(
            zip(times, roof_ys, wheel_ys, rel_displacements)):
        row_fill = fill_even if i % 2 == 0 else fill_odd
        row = [
            i,
            round(t, 4),
            round(ry, 2) if not np.isnan(ry) else None,
            round(wy, 2) if not np.isnan(wy) else None,
            None,   # roof_x placeholder (extend if you track X too)
            None,   # wheel_x placeholder
            round(rd, 2) if not np.isnan(rd) else None,
        ]
        ws.append(row)
        for cell in ws[i + 2]:
            cell.fill   = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Chart ────────────────────────────────────────────────
    wc = wb.create_sheet("Chart")
    n_rows = len(times) + 1   # +1 for header

    # --- Chart A: absolute positions ---
    chart_pos = LineChart()
    chart_pos.title   = "Roof & Wheel Vertical Position"
    chart_pos.style   = 10
    chart_pos.y_axis.title = f"Y position ({unit})"
    chart_pos.x_axis.title = "Frame"
    chart_pos.height  = 12
    chart_pos.width   = 24

    roof_ref  = Reference(ws, min_col=3, min_row=1, max_row=n_rows)
    wheel_ref = Reference(ws, min_col=4, min_row=1, max_row=n_rows)
    chart_pos.add_data(roof_ref,  titles_from_data=True)
    chart_pos.add_data(wheel_ref, titles_from_data=True)

    # Colour the series
    chart_pos.series[0].graphicalProperties.line.solidFill = "27AE60"  # green → roof
    chart_pos.series[1].graphicalProperties.line.solidFill = "E67E22"  # orange → wheel
    chart_pos.series[0].graphicalProperties.line.width = 15000   # 1.5 pt (EMUs)
    chart_pos.series[1].graphicalProperties.line.width = 15000

    wc.add_chart(chart_pos, "A1")

    # --- Chart B: suspension travel ---
    chart_susp = LineChart()
    chart_susp.title   = "Suspension Travel (wheel relative to roof)"
    chart_susp.style   = 10
    chart_susp.y_axis.title = f"Travel ({unit})   + droop  /  − compression"
    chart_susp.x_axis.title = "Frame"
    chart_susp.height  = 12
    chart_susp.width   = 24

    susp_ref = Reference(ws, min_col=7, min_row=1, max_row=n_rows)
    chart_susp.add_data(susp_ref, titles_from_data=True)
    chart_susp.series[0].graphicalProperties.line.solidFill = "2C3E50"
    chart_susp.series[0].graphicalProperties.line.width = 18000   # 1.8 pt

    wc.add_chart(chart_susp, "A23")

    wb.save(OUTPUT_XLSX)
    print(f"Excel file saved       → {OUTPUT_XLSX}")

    # -------------------------------------------------------------------
    # Plot
    # -------------------------------------------------------------------
    times_arr = np.array(times)
    rel_arr   = np.array(rel_displacements, dtype=float)
    roof_arr  = np.array(roof_ys, dtype=float)
    wheel_arr = np.array(wheel_ys, dtype=float)

    fig, axes = plt.subplots(2, 1, figsize=(12, 7), sharex=True)
    fig.suptitle("Car Suspension Analysis — Wheel Travel Relative to Roof",
                 fontsize=13, fontweight="bold")

    # Top: absolute vertical positions
    ax1 = axes[0]
    ax1.plot(times_arr, roof_arr,  color="#27ae60", label="Roof centre Y")
    ax1.plot(times_arr, wheel_arr, color="#e67e22", label="Wheel centre Y")
    ax1.set_ylabel(f"Vertical position ({unit})\n↓ positive = lower in frame")
    ax1.legend(loc="upper right")
    ax1.grid(True, alpha=0.3)
    ax1.invert_yaxis()   # image coords: 0 is top, so invert for intuition

    # Bottom: relative displacement
    ax2 = axes[1]
    ax2.axhline(0, color="grey", linewidth=0.8, linestyle="--")
    ax2.fill_between(times_arr, rel_arr, 0,
                     where=(rel_arr >= 0), alpha=0.25, color="#e74c3c",
                     label="Wheel extends (droop)")
    ax2.fill_between(times_arr, rel_arr, 0,
                     where=(rel_arr < 0), alpha=0.25, color="#3498db",
                     label="Wheel compresses (bump)")
    ax2.plot(times_arr, rel_arr, color="#2c3e50", linewidth=1.2)
    ax2.set_ylabel(f"Suspension travel ({unit})\n+ = droop  – = compression")
    ax2.set_xlabel("Time (s)")
    ax2.legend(loc="upper right")
    ax2.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(OUTPUT_PLOT, dpi=150)
    print(f"Plot saved             → {OUTPUT_PLOT}")
    plt.show()

    # Summary stats
    valid = rel_arr[~np.isnan(rel_arr)]
    if len(valid):
        print(f"\n--- Summary ({unit}) ---")
        print(f"  Max compression : {valid.min():.2f}")
        print(f"  Max droop       : {valid.max():.2f}")
        print(f"  Total travel    : {valid.max() - valid.min():.2f}")
        print(f"  Std deviation   : {valid.std():.2f}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Track car roof & wheel to measure suspension travel."
    )
    parser.add_argument("--video", required=True,
                        help="Path to the input video file.")
    parser.add_argument("--px-per-mm", type=float, default=None,
                        help="Pixel-to-mm scale (optional). E.g. 5.2 means 5.2 px = 1 mm.")
    parser.add_argument("--tracker", default="CSRT",
                        choices=["CSRT","KCF","MIL","MOSSE"],
                        help="OpenCV tracker algorithm (default: CSRT).")
    args = parser.parse_args()

    TRACKER_TYPE  = args.tracker
    PIXEL_PER_MM  = args.px_per_mm

    main(args.video)
